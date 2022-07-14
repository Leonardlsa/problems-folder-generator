import os
import re

import openpyxl
from openpyxl import load_workbook
import xlrd

CT = ".\\错题反馈\\"
QK = ".\\情况反馈\\"
blacklist = ['易兆诚', '万铭萱', '王梓洁', '文煜涵', '徐潇瀚', '万铭萱 ', '任惠仪','万铭萱 线','许铭萱','万铭宣','许铭轩']
for date in os.listdir(CT):

    if os.path.isdir(CT + date):                #日期目录
        status = [False, False, False]          #是否随堂
        for subject in os.listdir(CT + date):   #对每个该日期下的文件

            if subject[0] == '~':               #名字筛选
                subject = subject[2:]
            if re.search("化学", subject):       #断定科目
                sub = 1
            if re.search("物理", subject):
                sub = 0
            if re.search("数学", subject):
                sub = 2
            status[sub] = True                  #移除随堂标记
            ST = False
            now = CT + date + '\\' + subject
            WB = xlrd.open_workbook(CT + date + '\\' + subject, encoding_override="utf-8")
            sheet = WB.sheets()[0]
            NAMES = sheet.col_values(0)

            for i in range(1, len(NAMES)):
                name = sheet.cell_value(i, 0)
                if name[0] == '~':
                    name = name[2:]
                if name[0] == '$':
                    name = name[1:]
                if name in blacklist:
                    continue
                STU = QK + name + '.xlsx'
                sh = xlrd.open_workbook(STU, encoding_override="utf-8").sheets()[0]
                workbook = load_workbook(filename=STU)
                sheet1 = workbook['Table 1']
                CV = sh.col_values(0)
                if sheet.cell_value(i, 1) == '未交' or name == '请假':
                    sheet1.cell(CV.index(float(date))+1, 5 + sub, "未交")
                    sheet1.cell(CV.index(float(date))+1, 8 + sub, "未交")
                    sheet1.cell(CV.index(float(date))+1, 11 + sub, "未交")
                else:
                    sheet1.cell(CV.index(float(date))+1, 5 + sub, "√")
                    sheet1.cell(CV.index(float(date))+1, 8+ sub, "√")
                    sheet1.cell(CV.index(float(date))+1, 11 + sub, "√")
                print(name + '    ' + str(date) + '    '+str(sub) + '    '+ "had done.")
                workbook.save(QK + name + '.xlsx')

        if not status[0]:
            sub = 0
            for allpeople in os.listdir(QK):
                if allpeople[0] == '~':
                    allpeople = allpeople[2:]
                if allpeople[0] == '$':
                    allpeople = allpeople[1:]

                STU = QK+allpeople

                sh = xlrd.open_workbook(STU, encoding_override="utf-8").sheets()[0]
                workbook = load_workbook(filename=STU)
                sheet1 = workbook['Table 1']
                CV = sh.col_values(0)

                sheet1.cell(CV.index(float(date))+1, 5 + sub, "随堂")
                sheet1.cell(CV.index(float(date))+1, 8 + sub, "随堂")
                sheet1.cell(CV.index(float(date))+1, 11 + sub, "随堂")

                print(allpeople + '    ' + str(date) + '    ' + str(sub) + '    ' + "had done.")
                workbook.save(STU)
            status[sub] = True
        if not status[1]:
            sub = 1
            for allpeople in os.listdir(QK):
                if allpeople[0] == '~':
                    allpeople = allpeople[2:]
                if allpeople[0] == '$':
                    allpeople = allpeople[1:]
                STU = QK+allpeople
                sh = xlrd.open_workbook(STU, encoding_override="utf-8").sheets()[0]
                workbook = load_workbook(filename=STU)
                sheet1 = workbook['Table 1']
                CV = sh.col_values(0)
                sheet1.cell(CV.index(float(date))+1, 5 + sub, "随堂")
                sheet1.cell(CV.index(float(date))+1, 8 + sub, "随堂")
                sheet1.cell(CV.index(float(date))+1, 11 + sub, "随堂")
                print(allpeople + '    ' + str(date) + '    ' + str(sub) + '    ' + "had done.")
                workbook.save(STU)
            status[sub] = True

        if not status[2]:
            sub = 2
            for allpeople in os.listdir(QK):
                if allpeople[0] == '~':
                    allpeople = allpeople[2:]
                if allpeople[0] == '$':
                    allpeople = allpeople[1:]
                STU = QK+allpeople
                sh = xlrd.open_workbook(STU, encoding_override="utf-8").sheets()[0]
                workbook = load_workbook(filename=STU)
                sheet1 = workbook['Table 1']
                CV = sh.col_values(0)
                sheet1.cell(CV.index(float(date))+1, 5 + sub, "随堂")
                sheet1.cell(CV.index(float(date))+1, 8 + sub, "随堂")
                sheet1.cell(CV.index(float(date))+1, 11 + sub, "随堂")
                print(allpeople + '    ' + str(date) + '    ' + str(sub) + '    ' + "had done.")
                workbook.save(STU)
            status[sub] = True

